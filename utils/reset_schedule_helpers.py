# ----------- reset_schedule_helpers.py -----------

import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import re
from datetime import datetime, date, time
from openpyxl.styles import NamedStyle
from sf_connector.service_connector import connect_to_tenant_snowflake
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import Workbook


def generate_reset_schedule_template() -> Workbook:
    """
    Build a blank Reset Schedule Excel template workbook.

    - Creates a sheet named 'RESET_SCHEDULE_TEMPLATE'
    - Writes the canonical headers in row 1
    - Leaves all data rows empty for the user to fill

    This is the official template for users to paste/enter reset schedule rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "RESET_SCHEDULE_TEMPLATE"

    headers = [
        "CHAIN_NAME",    # A
        "STORE_NUMBER",  # B
        "STORE_NAME",    # C
        "PHONE_NUMBER",  # D
        "CITY",          # E
        "ADDRESS",       # F
        "STATE",         # G
        "COUNTY",        # H
        "TEAM_LEAD",     # I
        "RESET_DATE",    # J
        "RESET_TIME",    # K
        "STATUS",        # L
        "NOTES",         # M
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    return wb

def format_reset_schedule(workbook):
    """
    Validate and format a Reset Schedule workbook for upload.

    Responsibilities:
    - Enforce canonical headers on the 'RESET_SCHEDULE_TEMPLATE' sheet.
    - Validate required fields in each row:
        CHAIN_NAME, STORE_NUMBER, STORE_NAME, ADDRESS, CITY, RESET_DATE, RESET_TIME
    - Validate STORE_NUMBER is numeric.
    - Validate RESET_DATE as a real date (mm/dd/yyyy or Excel date).
    - Validate RESET_TIME as a real time (e.g., '8:00 AM', '13:00').
    - Normalize CHAIN_NAME and STORE_NAME to uppercase.
    - Apply consistent Excel date formatting to RESET_DATE.

    Returns:
        workbook (Workbook): Formatted openpyxl workbook if validation passes.
        None: If validation fails (errors are written to Streamlit).
    """
    # ---- Ensure the expected sheet exists ----
    if "RESET_SCHEDULE_TEMPLATE" not in workbook.sheetnames:
        st.error("❌ Worksheet 'RESET_SCHEDULE_TEMPLATE' was not found in the uploaded file.")
        return None

    ws = workbook["RESET_SCHEDULE_TEMPLATE"]

    # ---- Canonical headers ----
    header_names = [
        "CHAIN_NAME",    # A
        "STORE_NUMBER",  # B
        "STORE_NAME",    # C
        "PHONE_NUMBER",  # D
        "CITY",          # E
        "ADDRESS",       # F
        "STATE",         # G
        "COUNTY",        # H
        "TEAM_LEAD",     # I
        "RESET_DATE",    # J
        "RESET_TIME",    # K
        "STATUS",        # L
        "NOTES",         # M
    ]
    for idx, header in enumerate(header_names, start=1):
        ws.cell(row=1, column=idx, value=header)

    errors: list[str] = []

    # ------------------------------------------------------------------
    # 1) Required columns: collect all row-level errors (with row numbers)
    # ------------------------------------------------------------------
    required_columns = {
        "A": "CHAIN_NAME",
        "B": "STORE_NUMBER",
        "C": "STORE_NAME",
        "E": "CITY",
        "F": "ADDRESS",
        "J": "RESET_DATE",
        "K": "RESET_TIME",
    }

    for col_letter, col_name in required_columns.items():
        col_idx = column_index_from_string(col_letter)

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            if value is None or (isinstance(value, str) and value.strip() == ""):
                errors.append(
                    f"Row {row_idx}: {col_name} (column {col_letter}) is required and cannot be blank."
                )

    # ------------------------------------------------------------------
    # 2) Additional field-level validation (STORE_NUMBER, dates, times)
    # ------------------------------------------------------------------
    # Date style for Excel reset date column
    date_style = NamedStyle(name="reset_date_format", number_format="mm/dd/yyyy")

    for row_idx in range(2, ws.max_row + 1):
        # STORE_NUMBER must be numeric if present
        store_num_cell = ws.cell(row=row_idx, column=2)  # B
        store_val = store_num_cell.value
        if store_val is not None and str(store_val).strip() != "":
            try:
                int(str(store_val).strip())
            except ValueError:
                errors.append(
                    f"Row {row_idx}: STORE_NUMBER must be numeric (column B). Got '{store_val}'."
                )

        # RESET_DATE must be a valid date
        reset_date_cell = ws.cell(row=row_idx, column=10)  # J
        rd_val = reset_date_cell.value
        if rd_val is not None and not (isinstance(rd_val, str) and rd_val.strip() == ""):
            parsed_date = None

            if isinstance(rd_val, (datetime, date)):
                parsed_date = rd_val.date() if isinstance(rd_val, datetime) else rd_val
            elif isinstance(rd_val, (int, float)):
                # Excel serial date – let Snowflake/pandas handle it later;
                # we just trust that it's a valid Excel date at this point.
                # No extra error here.
                pass
            elif isinstance(rd_val, str):
                try:
                    parsed_date = datetime.strptime(rd_val.strip(), "%m/%d/%Y").date()
                except ValueError:
                    errors.append(
                        f"Row {row_idx}: RESET_DATE '{rd_val}' is not a valid date in mm/dd/yyyy format."
                    )

            if parsed_date is not None:
                reset_date_cell.value = parsed_date
                reset_date_cell.style = date_style

        # RESET_TIME must be a valid time
        reset_time_cell = ws.cell(row=row_idx, column=11)  # K
        rt_val = reset_time_cell.value
        if rt_val is not None and not (isinstance(rt_val, str) and rt_val.strip() == ""):
            parsed_time = None

            if isinstance(rt_val, time):
                parsed_time = rt_val
            elif isinstance(rt_val, datetime):
                parsed_time = rt_val.time()
            elif isinstance(rt_val, str):
                txt = rt_val.strip()
                for fmt in ("%I:%M %p", "%H:%M", "%H:%M:%S"):
                    try:
                        parsed_time = datetime.strptime(txt, fmt).time()
                        break
                    except ValueError:
                        continue
                if parsed_time is None:
                    errors.append(
                        f"Row {row_idx}: RESET_TIME '{rt_val}' is not a recognized time "
                        "(expected formats like '8:00 AM' or '13:00')."
                    )

            if parsed_time is not None:
                reset_time_cell.value = parsed_time

        # Normalize CHAIN_NAME + STORE_NAME to uppercase
        chain_cell = ws.cell(row=row_idx, column=1)  # A
        if isinstance(chain_cell.value, str):
            chain_cell.value = chain_cell.value.strip().upper()

        store_name_cell = ws.cell(row=row_idx, column=3)  # C
        if isinstance(store_name_cell.value, str):
            store_name_cell.value = store_name_cell.value.strip().upper()

    # ------------------------------------------------------------------
    # 3) If any errors, bail out with messages; otherwise succeed
    # ------------------------------------------------------------------
    if errors:
        st.error("❌ Reset Schedule validation failed. Please fix the issues below and re-upload:")
        for msg in errors:
            st.markdown(f"- {msg}")
        return None

    st.success("✅ Reset Schedule template validated and formatted successfully.")
    return workbook



def upload_reset_data(df: pd.DataFrame, selected_chain: str):
    """
    Uploads reset schedule data to Snowflake after deleting existing entries for the selected chain.

    Args:
        df (pd.DataFrame): Formatted reset schedule data.
        selected_chain (str): Chain name to target for deletion + upload.
    """
    toml_info = st.session_state.get("toml_info")
    tenant_id = st.session_state.get("tenant_id")
    if not toml_info or not tenant_id:
        st.error("TOML or tenant ID missing from session state.")
        return

    conn = connect_to_tenant_snowflake(toml_info)
    selected_chain = selected_chain.upper()

    if df['CHAIN_NAME'].isnull().any() or df['STORE_NAME'].isnull().any():
        st.warning("CHAIN_NAME and STORE_NAME cannot be null. Please correct and try again.")
        return

    mismatches = df.loc[df['CHAIN_NAME'].str.upper() != selected_chain]
    if not mismatches.empty:
        st.warning(f"CHAIN_NAME mismatch: Found {len(mismatches)} rows not matching '{selected_chain}'.")
        return

    try:
        now = datetime.now()
        today = date.today()

        df['TENANT_ID'] = tenant_id

        # Convert to Snowflake-safe ISO strings (pyformat cannot bind Python datetime objects)
        df['CREATED_AT'] = now.strftime("%Y-%m-%d %H:%M:%S")
        df['UPDATED_AT'] = now.strftime("%Y-%m-%d %H:%M:%S")
        df['LAST_LOAD_DATE'] = today.strftime("%Y-%m-%d")

        # Normalize RESET_DATE → YYYY-MM-DD
        df['RESET_DATE'] = (
            pd.to_datetime(df['RESET_DATE'], errors='coerce')
              .dt.strftime("%Y-%m-%d")
        )

        # Normalize RESET_TIME → HH:MM:SS
        df['RESET_TIME'] = (
            pd.to_datetime(df['RESET_TIME'], errors='coerce')
              .dt.strftime("%H:%M:%S")
        )


        df.replace({np.nan: None, '': None}, inplace=True)

        expected_columns = [
            'CHAIN_NAME', 'STORE_NUMBER', 'STORE_NAME', 'PHONE_NUMBER', 'CITY', 'ADDRESS',
            'STATE', 'COUNTY', 'TEAM_LEAD', 'RESET_DATE', 'RESET_TIME', 'STATUS', 'NOTES',
            'TENANT_ID', 'CREATED_AT', 'UPDATED_AT', 'LAST_LOAD_DATE'
        ]

        df = df[expected_columns]

        placeholders = ', '.join(['%s'] * len(expected_columns))

        delete_query = f"DELETE FROM RESET_SCHEDULE WHERE CHAIN_NAME = '{selected_chain}'"
        insert_query = f"INSERT INTO RESET_SCHEDULE ({', '.join(expected_columns)}) VALUES ({placeholders})"

        with conn.cursor() as cur:
            cur.execute("BEGIN;")
            st.info(f"Removing existing RESET_SCHEDULE records for: {selected_chain}")
            cur.execute(delete_query)

            st.info("Inserting new records into RESET_SCHEDULE...")
            cur.executemany(insert_query, df.values.tolist())
            conn.commit()

        st.success(f"✅ Reset schedule uploaded for chain: {selected_chain}")

    except Exception as e:
        conn.rollback()
        st.error(f"❌ Upload failed: {e}")

    finally:
        conn.close()
