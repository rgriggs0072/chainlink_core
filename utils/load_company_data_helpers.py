# utils/load_company_data_helpers.py
"""
Load Company Data Helpers

Overview for future devs:
- Houses all formatting & upload helpers for:
    * SALES_REPORT
    * CUSTOMERS
    * PRODUCTS
    * SUPPLIER_COUNTY
- Contains both:
    * Legacy Excel-based formatters (openpyxl) kept for rollback
    * New DataFrame-based validation/template logic (starting with CUSTOMERS)
- All Snowflake writes use:
    * Tenant-aware connection via connect_to_tenant_snowflake
    * TRUNCATE + INSERT pattern wrapped in explicit transactions
    * Audit fields: TENANT_ID, CREATED_AT, UPDATED_AT, LAST_LOAD_DATE

Notes:
- This module assumes st.session_state["toml_info"] and ["tenant_id"] are set
  by the login / tenant bootstrap logic before uploads are called.
"""

from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook

from sf_connector.service_connector import connect_to_tenant_snowflake
from utils.class_validation_helpers import (
    ColumnRule,
    ValidationResult,
    validate_dataframe,
)

# =============================================================================
# 🔧 LEGACY FORMATTERS (Excel / openpyxl)
# =============================================================================

def remove_total_rows_worksheet(ws):
    """
    Remove any rows where column A contains the word 'TOTAL'
    (case-insensitive, trims whitespace).

    This operates in-place on an openpyxl worksheet.
    """
    rows_to_delete = []

    # Start at row 2 to avoid nuking the header row
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)  # column A
        val = cell.value
        if val is None:
            continue

        text = str(val).strip().upper()

        # Be forgiving: match anything that starts with or equals TOTAL
        if text == "TOTAL" or text.startswith("TOTAL"):
            rows_to_delete.append(row_idx)

    # Delete from bottom up so indices don't shift
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)





def format_sales_report(workbook):
    """
    Legacy formatter for the Sales Report export from Encompass.

    Expected source layout (row 1 headers):
      A: Chain Store Number
      B: Customer Name
      C: Shipping Address
      D: Salesman Assigned
      E: Product Name
      F: Carrier UPC
      G: Buyer Count <date-range>
      H: Buyer Count %

    Output:
      - Single sheet named 'SALES REPORT'
      - Columns (in order):
            STORE_NUMBER,
            STORE_NAME,
            ADDRESS,
            SALESPERSON,
            PRODUCT_NAME,
            UPC,
            PURCHASED_YES_NO

      - Column H (Buyer Count %) is removed.
      - UPC is normalized to digits-only (no hyphens, spaces, etc.).
      - PURCHASED_YES_NO is derived from Buyer Count:
            > 0  -> 1
            else -> 0

    This is a legacy helper kept for users who still pull the raw Encompass
    export instead of using the new template-based flow. The output is shaped
    to match SALES_SCHEMA so it can feed write_salesreport_to_snowflake().
    """
    try:
        # Prefer an existing "SALES REPORT" sheet if present; otherwise use the first sheet.
        sheet_name = (
            "SALES REPORT"
            if "SALES REPORT" in workbook.sheetnames
            else workbook.sheetnames[0]
        )

        # Drop all other sheets
        for name in list(workbook.sheetnames):
            if name != sheet_name:
                del workbook[name]

        ws = workbook[sheet_name]
        ws.title = "SALES REPORT"

        # --- Rename header row to match SALES_SCHEMA ---
        ws["A1"].value = "STORE_NUMBER"
        ws["B1"].value = "STORE_NAME"
        ws["C1"].value = "ADDRESS"
        ws["D1"].value = "SALESPERSON"
        ws["E1"].value = "PRODUCT_NAME"
        ws["F1"].value = "UPC"
        ws["G1"].value = "PURCHASED_YES_NO"

        # Drop column H (Buyer Count %) if present
        if ws.max_column >= 8:
            ws.delete_cols(8)

        # --- Normalize UPC: digits-only in column F ---
        for cell in ws["F"][2:]:  # skip header
            val = cell.value
            if val is None:
                continue
            s = str(val).strip()
            digits_only = "".join(ch for ch in s if ch.isdigit())
            cell.value = digits_only if digits_only else None

        # --- Derive PURCHASED_YES_NO from Buyer Count in column G ---
        for cell in ws["G"][2:]:  # skip header row
            raw = cell.value

            # Blank or non-numeric → treat as 0
            if raw is None or str(raw).strip() == "":
                cell.value = 0
                continue

            try:
                val = float(raw)
                cell.value = 1 if val > 0 else 0
            except Exception:
                cell.value = 0

        # --- Clean STORE_NAME (col B): drop " #XXXX" suffix, clean punctuation ---
        for cell in ws["B"][2:]:
            if isinstance(cell.value, str):
                raw = cell.value.strip()
                if "#" in raw:
                    raw = raw.split("#", 1)[0].strip()
                cleaned = (
                    raw.replace(",", " ")
                    .replace(" 's", "")
                    .replace("'", "")
                    .strip()
                )
                cell.value = cleaned

        # --- Clean other text columns (ADDRESS, SALESPERSON, PRODUCT_NAME) ---
        for col_letter in ["C", "D", "E"]:
            for cell in ws[col_letter][2:]:
                if isinstance(cell.value, str):
                    cell.value = (
                        cell.value.replace(",", " ")
                        .replace(" 's", "")
                        .replace("'", "")
                        .strip()
                    )

        # ------------------------------------------------------------------
        # FINAL SAFETY: rebuild the sheet via pandas and drop the TOTAL row
        # ------------------------------------------------------------------
        data = list(ws.values)
        if not data:
            return workbook

        header_row = data[0]
        body_rows = data[1:]

        # Build DataFrame from current sheet contents
        df = pd.DataFrame(body_rows, columns=header_row)

        # Remove any row where STORE_NUMBER is 'TOTAL' (case-insensitive)
        if "STORE_NUMBER" in df.columns:
            mask = ~df["STORE_NUMBER"].astype(str).str.strip().str.upper().eq("TOTAL")
            df = df[mask].reset_index(drop=True)

        # Clear the worksheet and write back the cleaned data
        ws.delete_rows(1, ws.max_row)

        # Write headers
        ws.append(list(df.columns))

        # Write rows
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))

        return workbook

    except Exception as e:
        st.error(f"Error formatting sales report: {str(e)}")
        return None



def format_customers_report(workbook):
    """
    Legacy formatter for the 'Customers' worksheet.

    Behavior (old app path):
    - Keeps only 'Customers' sheet
    - Removes auto-filter and extra sheets
    - Repositions and cleans store number & store name
    - Enforces numeric-only Store_Number
    - Writes a fixed header row matching legacy upload expectations

    This is kept for rollback/compat while the new validation flow
    (template-driven) is rolled out.
    """
    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name != "Customers":
                workbook.remove(workbook[sheet_name])

        ws = workbook["Customers"]
        ws.auto_filter.ref = None
        ws.insert_cols(3)

        # Strip '#XXXX' from store-name column (col D)
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                if "#" in str(cell.value):
                    ws.cell(row=cell.row, column=4).value = str(cell.value).split("#")[0]

        # Move store_number into col C
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            ws.cell(row=row[0].row, column=3).value = row[0].value

        ws.delete_cols(5)

        # Legacy header set
        headers = [
            "Customer_id",
            "Chain_Name",
            "Store_Number",
            "Store_Name",
            "Address",
            "City",
            "County",
            "Salesperson",
            "Account_Status",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Clean stray quotes
        for col in ["B", "E"]:
            for cell in ws[col]:
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.replace("'", "")

        # Validate that Store_Number is numeric
        invalid_rows = []
        for cell in ws["C"][1:]:
            if cell.value and not str(cell.value).isdigit():
                invalid_rows.append((cell.row, cell.value))

        if invalid_rows:
            st.error("Non-numeric Store Numbers found:")
            for row_num, val in invalid_rows:
                st.warning(f"Row {row_num}: '{val}'")
            st.stop()

        return workbook

    except Exception as e:
        st.error(f"Error formatting Customers sheet: {str(e)}")
        return None


# =============================================================================
# 🧪 COMMON NORMALIZATION HELPERS
# =============================================================================


def _normalize_str_series(series: pd.Series) -> pd.Series:
    """
    Normalize string columns:
    - Cast to str
    - Strip whitespace
    - Convert obvious 'nan'/'None' to real nulls
    """
    s = series.astype(str).str.strip()
    s = s.replace({"nan": None, "None": None, "": None})
    return s


def _normalize_store_name(name: str | None) -> str:
    """
    Normalize store names for comparison:
    - Treat None / NaN as empty string
    - Uppercase
    - Strip whitespace
    - Strip everything after a '#' (e.g. 'SMART & FINAL #829' -> 'SMART & FINAL')
    """
    if name is None:
        return ""
    s = str(name).upper().strip()
    if "#" in s:
        s = s.split("#", 1)[0].strip()
    return s


def _clean_upc(value):
    """
    Normalize UPC-like values by stripping all non-digit characters.

    Rules:
    - If value is int: use the integer as-is.
    - If value is float: convert to int first (removes .0 and sci-notation issues).
    - If value is string: strip, handle 'nan'/'None', and keep digits only.
    - Return cleaned digit string or None.
    """
    # Null / NaN handling
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        # Non-pandas scalars just fall through
        pass

    # Handle pure integers
    if isinstance(value, int):
        s = str(value)

    # Handle floats (Excel numbers come in as floats)
    elif isinstance(value, float):
        try:
            # Safely convert float → int to avoid .0 or sci-notation leakage
            s = str(int(round(value)))
        except Exception:
            s = str(value).strip()

    # Handle everything else as a string
    else:
        s = str(value).strip()

    if not s or s.lower() in ("nan", "none"):
        return None

    # Digits only
    digits_only = "".join(ch for ch in s if ch.isdigit())
    return digits_only or None


# =============================================================================
# 🧪 NEW CUSTOMERS TEMPLATE + VALIDATION LAYER
# =============================================================================

# Schema must match CUSTOMERS table (minus server-side audit cols).
CUSTOMERS_SCHEMA = [
    ColumnRule("CUSTOMER_ID", required=True, dtype="int"),
    ColumnRule("CHAIN_NAME", required=True, dtype="str"),
    ColumnRule("STORE_NUMBER", required=True, dtype="int"),
    ColumnRule("STORE_NAME", required=True, dtype="str"),
    ColumnRule("ADDRESS", required=True, dtype="str"),
    ColumnRule("CITY", required=True, dtype="str"),
    ColumnRule("COUNTY", required=True, dtype="str"),
    ColumnRule("SALESPERSON", required=True, dtype="str", allow_blank=False),
    ColumnRule("ACCOUNT_STATUS", required=True, dtype="str"),
    # TENANT_ID, CREATED_AT, UPDATED_AT, LAST_LOAD_DATE handled server-side.
]


def generate_customers_template() -> pd.DataFrame:
    """
    Create an empty Customers template DataFrame matching CUSTOMERS_SCHEMA.

    This is what users should paste into from Encompass (or other sources).
    """
    cols = [rule.name for rule in CUSTOMERS_SCHEMA]
    return pd.DataFrame(columns=cols)


def format_customers_upload(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Light normalization for Customers uploads using the OFFICIAL template.

    Rules:
    - Columns must match template names (case-insensitive, spaces/underscores ignored).
    - Normalize column names into canonical uppercase-underscore form.
    - Normalize key text columns (strip, handle nan/None).
    """
    df = raw_df.copy()

    def _norm(name: str) -> str:
        return name.strip().lower().replace(" ", "").replace("_", "")

    incoming_cols = list(df.columns)
    norm_map = {_norm(c): c for c in incoming_cols}

    target_cols = [rule.name for rule in CUSTOMERS_SCHEMA]

    rename = {}
    for target in target_cols:
        norm_target = _norm(target)
        if norm_target in norm_map:
            # Map whatever the user had (Customer_id, customer id, etc.) to our canonical name
            rename[norm_map[norm_target]] = target

    df = df.rename(columns=rename)

    # After renaming, clean up headers
    df.columns = [c.strip() for c in df.columns]

    # Normalize strings on key text columns
    for col in [
        "CHAIN_NAME",
        "STORE_NAME",
        "ADDRESS",
        "CITY",
        "COUNTY",
        "SALESPERSON",
        "ACCOUNT_STATUS",
    ]:
        if col in df.columns:
            df[col] = _normalize_str_series(df[col])

    return df


def validate_customers_upload(df: pd.DataFrame) -> ValidationResult:
    """
    Apply schema-based validation to Customers upload:
    - Required columns
    - Types (int/str)
    - Non-blank required fields
    - Drops & warns on extra columns
    """
    return validate_dataframe(df, CUSTOMERS_SCHEMA)


def validate_customers_against_existing_chains(
    df: pd.DataFrame,
) -> tuple[list[str], list[str]]:
    """
    Cross-validate CHAIN_NAME values in the upload against existing CUSTOMERS
    for this tenant, to catch issues like 'FOODMAXX' vs 'FOOD MAXX'.

    Returns:
        (errors, warnings)
    """
    conn, cursor, tenant_id = _get_conn_and_cursor()
    if not conn or not tenant_id:
        # If we don't have context, skip this check quietly.
        return [], []

    try:
        cursor.execute(
            "SELECT DISTINCT CHAIN_NAME FROM CUSTOMERS WHERE TENANT_ID = %s",
            (tenant_id,),
        )
        rows = cursor.fetchall()
    except Exception:
        # If query fails for any reason, don't block the upload on this cross-check
        return [], []
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    known_raw = [r[0] for r in rows if r[0] is not None]
    if not known_raw:
        # First load for this tenant: nothing to compare yet
        return [], []

    def norm_name(s: str) -> str:
        # Uppercase and remove spaces for normalization
        return "".join(str(s).upper().split())

    # Map from normalized -> original name
    known_norm_map = {norm_name(x): x for x in known_raw}

    upload_chains = df["CHAIN_NAME"].dropna().astype(str)
    upload_unique = sorted(upload_chains.unique())

    errors: list[str] = []
    warnings: list[str] = []

    for val in upload_unique:
        n = norm_name(val)
        if n not in known_norm_map:
            errors.append(
                f"Chain '{val}' does not match any existing CHAIN_NAME for this tenant."
            )
        else:
            existing = known_norm_map[n]
            if existing != val:
                warnings.append(
                    f"Chain '{val}' differs from existing '{existing}'. "
                    f"Consider standardizing to '{existing}' for consistency."
                )

    return errors, warnings


# =============================================================================
# 🧪 SALES REPORT TEMPLATE + VALIDATION LAYER
# =============================================================================


def _validate_sales_upc(series: pd.Series) -> list[str]:
    """
    Per-row UPC validator for SALES_REPORT uploads.

    Rules:
      - After cleaning (digits only), UPC must not be empty.
      - Length must be 10, 11 or 12 digits.
      - We report row numbers relative to a 1-based Excel-style sheet:
        assume header is row 1, so DataFrame index 0 => row 2, etc.
    """
    errors: list[str] = []

    for idx, raw in series.items():
        row_num = idx + 2  # header is row 1 in the typical Excel template
        cleaned = _clean_upc(raw)

        # Completely missing / no digits at all
        if cleaned is None:
            errors.append(
                f"Row {row_num}: UPC is missing or contains no usable digits (value={repr(raw)})."
            )
            continue

        # Must be all digits (paranoid check; _clean_upc already strips non-digits)
        if not cleaned.isdigit():
            errors.append(
                f"Row {row_num}: UPC '{raw}' cleaned to '{cleaned}', which still contains non-digit characters."
            )
            continue

        # Length check: 10, 11 or 12 digits
        if len(cleaned) not in (10, 11, 12):
            errors.append(
                f"Row {row_num}: UPC '{raw}' cleaned to '{cleaned}' has length {len(cleaned)}; "
                f"expected 10, 11 or 12 digits."
            )

    return errors


def _validate_purchased_flag(series: pd.Series) -> list[str]:
    """
    Validator for PURCHASED_YES_NO:
      - Must be 0 or 1 for every row.
    """
    errors: list[str] = []

    for idx, raw in series.items():
        row_num = idx + 2  # again, header = row 1
        if pd.isna(raw):
            errors.append(
                f"Row {row_num}: PURCHASED_YES_NO is blank; expected 0 or 1."
            )
            continue

        try:
            val = int(raw)
        except Exception:
            errors.append(
                f"Row {row_num}: PURCHASED_YES_NO '{raw}' is not an integer; expected 0 or 1."
            )
            continue

        if val not in (0, 1):
            errors.append(
                f"Row {row_num}: PURCHASED_YES_NO '{raw}' is not valid; expected 0 or 1."
            )

    return errors


# Schema must match SALES_REPORT table (minus audit/audit-like columns).
SALES_SCHEMA = [
    ColumnRule("STORE_NUMBER", required=True, dtype="int"),
    ColumnRule("STORE_NAME", required=True, dtype="str"),
    ColumnRule("ADDRESS", required=True, dtype="str"),
    ColumnRule("SALESPERSON", required=True, dtype="str"),
    ColumnRule("PRODUCT_NAME", required=True, dtype="str"),
    ColumnRule(
        "UPC",
        required=True,
        dtype="str",
        allow_blank=False,
        validators=[_validate_sales_upc],
    ),
    ColumnRule(
        "PURCHASED_YES_NO",
        required=True,
        dtype="int",
        allow_blank=False,
        validators=[_validate_purchased_flag],
    ),
    # TENANT_ID, CHAIN_NAME, CREATED_AT, LAST_LOAD_DATE are set server-side.
]


def generate_sales_template() -> pd.DataFrame:
    """
    Create an empty Sales Report template DataFrame matching SALES_SCHEMA.

    This is what users should paste into instead of raw Encompass exports.
    """
    cols = [rule.name for rule in SALES_SCHEMA]
    return pd.DataFrame(columns=cols)


def format_sales_upload(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize a Sales Report upload that uses the OFFICIAL template.

    Rules:
    - Columns can be case-insensitive and ignore spaces/underscores.
    - We rename headers into canonical names in SALES_SCHEMA.
    - We normalize:
        * STORE_NAME: strip '#XXXX' suffix, remove quotes, uppercase.
        * UPC: digits only via _clean_upc().
        * Text fields: trimmed, normalized via _normalize_str_series.
        * PURCHASED_YES_NO: int 0/1 (no 1.00 floats).
    """
    df = raw_df.copy()

    # Header normalization similar to Customers
    def _norm(name: str) -> str:
        return name.strip().lower().replace(" ", "").replace("_", "")

    incoming_cols = list(df.columns)
    norm_map = {_norm(c): c for c in incoming_cols}
    target_cols = [rule.name for rule in SALES_SCHEMA]

    rename = {}
    for target in target_cols:
        norm_target = _norm(target)
        if norm_target in norm_map:
            rename[norm_map[norm_target]] = target

    # Rename to canonical names where possible
    df = df.rename(columns=rename)

    # Clean headers
    df.columns = [c.strip() for c in df.columns]

    # Clean STORE_NAME if present (strip chain # suffix, upper, remove quotes)
    if "STORE_NAME" in df.columns:
        df["STORE_NAME"] = (
            df["STORE_NAME"]
            .astype(str)
            .str.upper()
            .str.replace("'", "", regex=False)
            .str.split("#")
            .str[0]
            .str.strip()
        )

    # Clean UPC if present
    if "UPC" in df.columns:
        df["UPC"] = df["UPC"].apply(_clean_upc)

    # Normalize basic text fields
    for col in ["ADDRESS", "SALESPERSON", "PRODUCT_NAME"]:
        if col in df.columns:
            df[col] = _normalize_str_series(df[col])

    # Ensure PURCHASED_YES_NO comes through as 0/1 ints (no 1.00 floats)
    if "PURCHASED_YES_NO" in df.columns:
        df["PURCHASED_YES_NO"] = (
            df["PURCHASED_YES_NO"].astype(float).fillna(0).round(0).astype(int)
        )

    return df


def validate_sales_upload(df: pd.DataFrame) -> ValidationResult:
    """
    Apply schema-based validation to Sales Report upload:
    - Required columns
    - Types (int/str)
    - Non-blank required fields
    - Drops & warns on extra columns
    """
    return validate_dataframe(df, SALES_SCHEMA)


def _build_missing_store_number_error(
    df: pd.DataFrame, missing_values: set[int]
) -> str:
    """
    Build a detailed message showing which Excel rows contain STORE_NUMBERs
    that do NOT exist in CUSTOMERS for this tenant.
    """
    lines = [
        "Some STORE_NUMBER values in your upload do not exist in the CUSTOMERS table for this tenant:",
        "Problem rows (Excel row numbers):",
        "",
    ]

    for idx, row in df.iterrows():
        try:
            sn = int(row["STORE_NUMBER"])
        except Exception:
            continue

        if sn in missing_values:
            excel_row = idx + 2  # header is row 1
            store_name = str(row.get("STORE_NAME", "")).strip()
            lines.append(
                f"• Row {excel_row}: STORE_NUMBER={sn}, STORE_NAME='{store_name}'"
            )

    lines.append("")
    lines.append(
        "To fix: correct these STORE_NUMBER values in your file or add these stores "
        "to the Customers table first, then re-upload."
    )

    return "\n".join(lines)


def validate_sales_against_customers(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    """
    Cross-validate STORE_NUMBER + STORE_NAME pairs in the Sales upload
    against existing CUSTOMERS for this tenant.

    Rules:
      - HARD ERROR if:
          * STORE_NUMBER does NOT exist in CUSTOMERS for this tenant, OR
          * STORE_NUMBER exists but STORE_NAME does NOT match ANY existing
            STORE_NAME for that STORE_NUMBER (after normalization).
      - OK if:
          * (STORE_NUMBER, STORE_NAME) matches at least one CUSTOMERS row
            (supports multiple chains sharing the same store_number).

    Returns:
        (errors, warnings)
    """
    conn, cursor, tenant_id = _get_conn_and_cursor()
    if not conn or not tenant_id:
        return [], []

    try:
        cursor.execute(
            """
            SELECT DISTINCT STORE_NUMBER, STORE_NAME
            FROM CUSTOMERS
            WHERE TENANT_ID = %s
              AND STORE_NUMBER IS NOT NULL;
            """,
            (tenant_id,),
        )
        rows = cursor.fetchall()
    except Exception:
        # Don't block upload if the cross-check fails
        return [], []
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    if not rows:
        # No customer data yet for this tenant; skip this check
        return [], []

    def _norm_name(name: str) -> str:
        if name is None:
            return ""
        s = str(name).upper().strip()
        # Strip trailing "#xxxx" patterns used in some exports
        if "#" in s:
            s = s.split("#")[0].strip()
        return s

    # For each STORE_NUMBER, keep *all* normalized + raw names
    store_to_names_norm: dict[int, set[str]] = {}
    store_to_names_raw: dict[int, set[str]] = {}

    for sn, sname in rows:
        if sn is None:
            continue
        try:
            sn_int = int(sn)
        except Exception:
            continue

        norm = _norm_name(sname)
        raw = str(sname) if sname is not None else ""

        store_to_names_norm.setdefault(sn_int, set()).add(norm)
        store_to_names_raw.setdefault(sn_int, set()).add(raw)

    errors: list[str] = []
    warnings: list[str] = []

    if "STORE_NUMBER" not in df.columns or "STORE_NAME" not in df.columns:
        errors.append(
            "STORE_NUMBER and/or STORE_NAME columns are missing; "
            "cannot cross-check against CUSTOMERS."
        )
        return errors, warnings

    problem_rows: list[str] = []

    # Walk each uploaded row and compare against CUSTOMER lookup
    for idx, row in df.reset_index(drop=True).iterrows():
        raw_sn = row.get("STORE_NUMBER")
        raw_name = row.get("STORE_NAME")
        excel_row = idx + 2  # header is Excel row 1

        if pd.isna(raw_sn):
            continue

        try:
            sn_int = int(raw_sn)
        except Exception:
            # If it's totally non-numeric, schema validation should catch it
            continue

        upload_name_norm = _norm_name(raw_name)

        # 1) STORE_NUMBER not in CUSTOMERS at all -> HARD ERROR
        if sn_int not in store_to_names_norm:
            problem_rows.append(
                f"Row {excel_row}: STORE_NUMBER={sn_int}, STORE_NAME='{raw_name}' "
                f"does not exist in CUSTOMERS for this tenant."
            )
            continue

        # 2) STORE_NUMBER exists, but STORE_NAME must match at least one entry
        known_norms = store_to_names_norm.get(sn_int, set())
        known_raws = sorted(store_to_names_raw.get(sn_int, set()))

        if upload_name_norm not in known_norms:
            expected_list = ", ".join(f"'{n}'" for n in known_raws if n)
            problem_rows.append(
                f"Row {excel_row}: STORE_NUMBER={sn_int}, STORE_NAME='{raw_name}' "
                f"does not match any existing Customers records for this store. "
                f"Expected one of: {expected_list}."
            )

    if problem_rows:
        lines = [
            "Some STORE_NUMBER / STORE_NAME combinations in your upload do not match the CUSTOMERS table for this tenant.",
            "This is a hard validation error; please correct your file before re-uploading.",
            "Problem rows (Excel row numbers):",
        ]
        lines.extend(f"• {msg}" for msg in problem_rows)
        errors.append("\n".join(lines))

    return errors, warnings


# =============================================================================
# 📌 SUPPLIER BY COUNTY – REQUIRED COLUMNS (Upload Template)
# =============================================================================

SUPPLIER_COUNTY_REQUIRED_COLUMNS = [
    "SUPPLIER",
    "COUNTY",
    "STATUS",
]


def generate_supplier_county_template() -> pd.DataFrame:
    """
    Generates an empty Supplier by County template.
    Tenant fields are added during upload and not included here.
    """
    return pd.DataFrame(columns=SUPPLIER_COUNTY_REQUIRED_COLUMNS)


def validate_supplier_county_upload(
    df_raw: pd.DataFrame,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Validate uploaded Supplier by County template.

    Requirements:
    - SUPPLIER: required, non-empty
    - COUNTY: required, non-empty
    - STATUS: required ("Yes" or "No"; case-insensitive)
    """
    errors: list[str] = []

    df = df_raw.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    # Check columns
    missing = [c for c in SUPPLIER_COUNTY_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        errors.append("Missing required columns: " + ", ".join(missing))
        return df, errors

    # Restrict to required columns only
    df = df[SUPPLIER_COUNTY_REQUIRED_COLUMNS].copy()

    # Drop empty rows
    df = df.dropna(how="all", subset=SUPPLIER_COUNTY_REQUIRED_COLUMNS)

    # Clean values
    for col in SUPPLIER_COUNTY_REQUIRED_COLUMNS:
        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .replace({"nan": "", "NaN": "", "NONE": "", "None": ""})
        )

    # Validate SUPPLIER
    missing_supplier = df["SUPPLIER"] == ""
    if missing_supplier.any():
        rows = [i + 2 for i in df.index[missing_supplier]]
        errors.append("SUPPLIER is required for row(s): " + ", ".join(map(str, rows)))

    # Validate COUNTY
    missing_county = df["COUNTY"] == ""
    if missing_county.any():
        rows = [i + 2 for i in df.index[missing_county]]
        errors.append("COUNTY is required for row(s): " + ", ".join(map(str, rows)))

    # Validate STATUS
    allowed = {"YES", "NO"}
    bad_status_mask = ~df["STATUS"].str.upper().isin(allowed)

    if bad_status_mask.any():
        rows = [i + 2 for i in df.index[bad_status_mask]]
        bad_vals = df.loc[bad_status_mask, "STATUS"].unique().tolist()
        errors.append(
            f"Invalid STATUS values {bad_vals}. Only 'Yes' or 'No' allowed. Rows: "
            + ", ".join(map(str, rows))
        )

    return df, errors


def format_supplier_by_county(file_content) -> pd.DataFrame:
    """
    Formats the uploaded Supplier by County pivot table Excel file into a normalized DataFrame
    ready for upload to the Snowflake SUPPLIER_COUNTY table.

    - Expects a sheet named 'Report'
    - Drops 'TOTAL' column if present
    - Renames 'Supplier / County' to 'Supplier'
    - Unpivots county columns into a single 'County' column
    - Converts values (1 → 'Yes', NaN → 'No')
    """
    try:
        xls = pd.ExcelFile(file_content)

        if "Report" not in xls.sheet_names:
            st.error("❌ Sheet named 'Report' not found in the Excel file.")
            st.info(
                "Please rename the sheet you want formatted to 'Report' and try again."
            )
            return None

        df = xls.parse("Report")

        if "Supplier / County" not in df.columns:
            st.error("❌ Column 'Supplier / County' not found in 'Report' sheet.")
            return None

        if "TOTAL" in df.columns:
            df = df.drop(columns=["TOTAL"])

        df.rename(columns={"Supplier / County": "Supplier"}, inplace=True)

        df_melted = pd.melt(
            df, id_vars=["Supplier"], var_name="County", value_name="Status"
        )

        df_melted["Status"] = df_melted["Status"].apply(
            lambda x: "Yes" if x == 1 else "No" if pd.isna(x) else str(x)
        )

        st.success("✅ Supplier by County formatting complete.")
        return df_melted

    except Exception as e:
        st.error(f"❌ Failed to format Supplier by County report: {str(e)}")
        return None


# ------------------------------------------------------------------------------------------------------------------------
# New Products template + validation layer (legacy openpyxl formatter kept for rollback)
# ------------------------------------------------------------------------------------------------------------------------

# =============================================================================
# 📄 PRODUCTS TEMPLATE (for download)
# =============================================================================


def create_products_template_workbook():
    """
    Build an in-memory Excel workbook for the Products template.

    Columns:
      - PRODUCT_ID       (optional numeric)
      - SUPPLIER        (required)
      - PRODUCT_NAME    (required)
      - PACKAGE         (required)
      - CARRIER_UPC     (required, digits only, <= 20 chars)
      - PRODUCT_MANAGER (optional)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "PRODUCT_ID",
        "SUPPLIER",
        "PRODUCT_NAME",
        "PACKAGE",
        "CARRIER_UPC",
        "PRODUCT_MANAGER",
    ]
    ws.append(headers)

    return wb


# =============================================================================
# 📦 PRODUCTS FORMATTER (legacy openpyxl path)
# =============================================================================


def format_product_workbook(workbook):
    """
    Cleans and normalizes the 'Products' worksheet for upload.

    - Moves, cleans, and renames columns
    - Removes commas, apostrophes, hyphens
    - Sets PRODUCT_MANAGER column and removes unused columns
    """
    try:
        ws = workbook["Products"]

        # Move col G into col B
        col_g_data = [cell.value for cell in ws["G"]]
        ws.insert_cols(2)
        for cell, value in zip(ws["B"], col_g_data):
            cell.value = value
        for cell in ws["G"]:
            cell.value = None

        # Move col E into col D and drop original
        col_e_data = [cell.value for cell in ws["E"]]
        for cell, value in zip(ws["D"], col_e_data):
            cell.value = value
        ws.delete_cols(5)

        # Strip hyphens from carrier UPC column (E)
        for cell in ws["E"]:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("-", "")

        # Clean all string cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace(",", "").replace("'", "")

        # PRODUCT_MANAGER column
        ws["F1"].value = "PRODUCT_MANAGER"
        for cell in ws["F"][1:]:
            cell.value = None

        # Drop any columns beyond F
        if ws.max_column >= 7:
            ws.delete_cols(7)

        # Fill missing UPC with placeholder numeric
        for cell in ws["E"][1:]:
            if cell.value is None:
                cell.value = 999999999999

        return workbook

    except Exception as e:
        st.error(f"❌ Error formatting product data: {str(e)}")
        return None


# =============================================================================
# 📤 DOWNLOAD HELPER
# =============================================================================


def download_workbook(workbook, filename: str) -> None:
    """
    Stream an openpyxl workbook to the user as an .xlsx download.
    """
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    st.download_button(
        label="Download formatted file",
        data=stream.read(),
        file_name=filename,
        mime="application/vnd.ms-excel",
    )


# =============================================================================
# 🔐 SNOWFLAKE CONNECTION + TRANSACTION HELPERS
# =============================================================================


def _get_conn_and_cursor():
    """
    Open a tenant-scoped Snowflake connection and return (conn, cursor, tenant_id).

    Requires:
        st.session_state["toml_info"]
        st.session_state["tenant_id"]
    """
    toml_info = st.session_state.get("toml_info")
    tenant_id = st.session_state.get("tenant_id")
    if not toml_info or not tenant_id:
        st.error("❌ Missing tenant configuration.")
        return None, None, None

    conn = connect_to_tenant_snowflake(toml_info)
    if not conn:
        st.error("❌ Failed to connect to Snowflake.")
        return None, None, None

    return conn, conn.cursor(), tenant_id


def _finalize_transaction(cursor, conn, success_msg: str) -> None:
    """
    Commit transaction, close cursor + connection, and show success message.
    """
    conn.commit()
    cursor.close()
    conn.close()
    st.success(success_msg)


def _rollback_transaction(conn, cursor) -> None:
    """
    Rollback transaction and cleanly close cursor + connection.
    """
    if conn:
        conn.rollback()
    if cursor:
        cursor.close()
    if conn:
        conn.close()


def _build_audit_fields() -> tuple[str, str]:
    """
    Build standard audit fields:
        - now_ts: current timestamp as string
        - today_date: current date as YYYY-MM-DD
    """
    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    today_date = datetime.today().strftime("%Y-%m-%d")
    return now_ts, today_date


# =============================================================================
# 🔍 SALES ENRICHMENT HELPERS
# =============================================================================


def _fetch_customer_chain_lookup(conn, tenant_id: str) -> dict:
    """
    Pulls STORE_NUMBER, STORE_NAME, CHAIN_NAME from CUSTOMERS for this tenant.

    Normalizes STORE_NAME for consistency (upper, strip, '#xxx' removed).
    Returns:
        dict[int, str]: {store_number: chain_name}
    """
    lookup: dict[int, str] = {}

    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT STORE_NUMBER, STORE_NAME, CHAIN_NAME
                FROM CUSTOMERS
                WHERE TENANT_ID = %s;
                """,
                (tenant_id,),
            )
            rows = cur.fetchall()

        for store_number, store_name, chain_name in rows:
            if store_number is None:
                continue

            norm_name = str(store_name or "").upper().strip()
            norm_name = norm_name.split("#")[0].strip()

            lookup[int(store_number)] = chain_name

        return lookup

    except Exception as e:
        st.warning(f"⚠️ Could not enrich CHAIN_NAME from CUSTOMERS: {e}")
        return {}


# =============================================================================
# 📥 SALES_REPORT UPLOAD
# =============================================================================


def write_salesreport_to_snowflake(df: pd.DataFrame):
    """
    Upload cleaned sales report to Snowflake.

    Enhancements:
    - Cleans UPC (digits only, handles floats/ints/strings)
    - Cleans STORE_NAME ('SMART & FINAL #829' -> 'SMART & FINAL')
    - Enriches CHAIN_NAME from CUSTOMERS table per tenant (if column exists)
    - Uses TRUNCATE + INSERT wrapped in a transaction
    - Adds audit fields (TENANT_ID, CREATED_AT, UPDATED_AT, LAST_LOAD_DATE)

    NOTE:
    - This assumes SALES_REPORT has a CHAIN_NAME column. If it does not yet,
      either add it or remove CHAIN_NAME from the INSERT + records tuple.
    """
    df = df.copy()

    # Clean UPC
    if "UPC" in df.columns:
        df["UPC"] = df["UPC"].apply(_clean_upc)

    # Clean store name (strip '#xxxx', upper, strip quotes)
    if "STORE_NAME" in df.columns:
        df["STORE_NAME"] = (
            df["STORE_NAME"]
            .astype(str)
            .str.upper()
            .str.replace("'", "", regex=False)
            .str.split("#")
            .str[0]
            .str.strip()
        )

    # Ensure PURCHASED_YES_NO is int 0/1 (no floats)
    if "PURCHASED_YES_NO" in df.columns:
        df["PURCHASED_YES_NO"] = (
            df["PURCHASED_YES_NO"].astype(float).fillna(0).round(0).astype(int)
        )

    conn, cursor, tenant_id = _get_conn_and_cursor()
    if not conn:
        return

    now_ts, today_date = _build_audit_fields()

    # Build CHAIN_NAME enrichment map
    chain_lookup = _fetch_customer_chain_lookup(conn, tenant_id)

    # Add CHAIN_NAME column based on STORE_NUMBER
    df["CHAIN_NAME"] = df["STORE_NUMBER"].apply(
        lambda sn: chain_lookup.get(int(sn)) if pd.notna(sn) else None
    )

    try:
        cursor.execute("BEGIN;")
        cursor.execute("TRUNCATE TABLE SALES_REPORT;")

        records = [
            (
                row.STORE_NUMBER,
                row.STORE_NAME,
                row.ADDRESS,
                row.SALESPERSON,
                row.PRODUCT_NAME,
                row.UPC,
                row.PURCHASED_YES_NO,
                tenant_id,
                now_ts,  # CREATED_AT
                None,  # SALE_DATE (NULL)
                row.CHAIN_NAME,
                today_date,  # LAST_LOAD_DATE
            )
            for row in df.itertuples(index=False)
        ]

        insert_sql = """
            INSERT INTO SALES_REPORT (
                STORE_NUMBER,
                STORE_NAME,
                ADDRESS,
                SALESPERSON,
                PRODUCT_NAME,
                UPC,
                PURCHASED_YES_NO,
                TENANT_ID,
                CREATED_AT,
                SALE_DATE,
                CHAIN_NAME,
                LAST_LOAD_DATE
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        cursor.executemany(insert_sql, records)
        _finalize_transaction(cursor, conn, "✅ Sales report uploaded.")

    except Exception as e:
        _rollback_transaction(conn, cursor)
        st.error(f"❌ Sales report upload failed: {str(e)}")


# =============================================================================
# 📥 CUSTOMERS UPLOAD
# =============================================================================


def write_customers_to_snowflake(df: pd.DataFrame):
    """
    Upload validated customer data to Snowflake with full audit fields.

    - Uppercases all string-like values
    - Uses TRUNCATE + INSERT in a transaction
    """
    df = df.copy()
    df.fillna("NULL", inplace=True)
    df = df.applymap(lambda x: str(x).strip().upper() if pd.notna(x) else x)

    conn, cursor, tenant_id = _get_conn_and_cursor()
    if not conn:
        return

    now_ts, today_date = _build_audit_fields()

    try:
        cursor.execute("BEGIN;")
        cursor.execute("TRUNCATE TABLE CUSTOMERS;")

        records = [
            (
                row.CUSTOMER_ID,
                row.CHAIN_NAME,
                row.STORE_NUMBER,
                row.STORE_NAME,
                row.ADDRESS,
                row.CITY,
                row.COUNTY,
                row.SALESPERSON,
                row.ACCOUNT_STATUS,
                tenant_id,
                now_ts,
                now_ts,
                today_date,
            )
            for row in df.itertuples(index=False)
        ]

        insert_sql = """
            INSERT INTO CUSTOMERS (
                CUSTOMER_ID,
                CHAIN_NAME,
                STORE_NUMBER,
                STORE_NAME,
                ADDRESS,
                CITY,
                COUNTY,
                SALESPERSON,
                ACCOUNT_STATUS,
                TENANT_ID,
                CREATED_AT,
                UPDATED_AT,
                LAST_LOAD_DATE
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        cursor.executemany(insert_sql, records)
        _finalize_transaction(cursor, conn, "✅ Customers uploaded.")

    except Exception as e:
        _rollback_transaction(conn, cursor)
        st.error(f"❌ Customer upload failed: {str(e)}")


# =============================================================================
# ✅ PRODUCTS TEMPLATE (DataFrame-based)
# =============================================================================


REQUIRED_PRODUCT_COLUMNS = [
    "PRODUCT_ID",
    "SUPPLIER",
    "PRODUCT_NAME",
    "PACKAGE",
    "CARRIER_UPC",
    "PRODUCT_MANAGER",
]


def generate_products_template() -> pd.DataFrame:
    """
    Build an in-memory Products template as a DataFrame.

    Columns:
      - PRODUCT_ID       (optional numeric)
      - SUPPLIER         (required)
      - PRODUCT_NAME     (required)
      - PACKAGE          (required)
      - CARRIER_UPC      (required, digits only, <= 20 chars)
      - PRODUCT_MANAGER  (optional)
    """
    cols = [
        "PRODUCT_ID",
        "SUPPLIER",
        "PRODUCT_NAME",
        "PACKAGE",
        "CARRIER_UPC",
        "PRODUCT_MANAGER",
    ]
    template_df = pd.DataFrame(columns=cols)
    return template_df


def validate_products_upload(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    Validate and clean the uploaded Products DataFrame.

    Rules:
    - Required columns:
        PRODUCT_ID, SUPPLIER, PRODUCT_NAME, PACKAGE, CARRIER_UPC, PRODUCT_MANAGER
    - PRODUCT_ID:
        * REQUIRED (no blank / NaN / 'nan' / whitespace)
        * Must be numeric (digits only)
        * Must be UNIQUE across the uploaded file
    - CARRIER_UPC:
        * Cleaned via _clean_upc (handles floats/ints/strings, strips non-digits)
        * REQUIRED (no blank)
        * Digits-only
        * Max length 20
        * Reject legacy placeholder '999999999999'
    """
    errors: list[str] = []

    # 1) Normalize column names immediately
    df = df_raw.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    # 2) Hard required-column check BEFORE anything else
    missing = [c for c in REQUIRED_PRODUCT_COLUMNS if c not in df.columns]
    if missing:
        errors.append(
            "Missing required columns in Products file: "
            + ", ".join(missing)
            + ". Please fix the formatted file/template and re-upload."
        )
        return df, errors

    # 3) Restrict to the columns we care about
    df = df[REQUIRED_PRODUCT_COLUMNS].copy()

    # 4) Drop completely-empty rows across required columns
    df = df.dropna(how="all", subset=REQUIRED_PRODUCT_COLUMNS)

    # 5) Clean string columns (except PRODUCT_ID – handled separately)
    str_cols = ["SUPPLIER", "PRODUCT_NAME", "PACKAGE", "CARRIER_UPC", "PRODUCT_MANAGER"]
    for col in str_cols:
        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .replace(
                {
                    "nan": "",
                    "NaN": "",
                    "NONE": "",
                    "None": "",
                }
            )
        )

    # ------------------------------------------------------------------
    # 6) PRODUCT_ID — REQUIRED, NUMERIC, UNIQUE
    # ------------------------------------------------------------------
    pid_col = df["PRODUCT_ID"]

    # a) detect missing BEFORE any numeric conversion
    missing_id_mask = pid_col.isna() | pid_col.astype(str).str.strip().isin(
        ["", "nan", "NaN", "NONE", "None"]
    )
    if missing_id_mask.any():
        idxs = df.index[missing_id_mask].tolist()
        human_rows = [i + 2 for i in idxs]  # Excel-style row numbers
        errors.append(
            "PRODUCT_ID is required and cannot be blank for row(s): "
            + ", ".join(map(str, human_rows))
            + "."
        )

    # b) enforce numeric: digits only at the string level
    pid_str = pid_col.astype(str).str.strip()
    bad_digit_mask = ~pid_str.str.match(r"^\d+$")
    bad_digit_mask = bad_digit_mask & ~missing_id_mask
    if bad_digit_mask.any():
        idxs = df.index[bad_digit_mask].tolist()
        human_rows = [i + 2 for i in idxs]
        errors.append(
            "PRODUCT_ID must be numeric (digits only) for row(s): "
            + ", ".join(map(str, human_rows))
            + "."
        )

    # c) uniqueness among valid (non-missing, numeric) IDs
    valid_mask = ~(missing_id_mask | bad_digit_mask)
    valid_pid_str = pid_str.where(valid_mask)
    dup_mask = valid_pid_str.duplicated(keep=False) & valid_mask
    if dup_mask.any():
        dup_ids = sorted(valid_pid_str[dup_mask].unique().tolist())
        errors.append(
            "Duplicate PRODUCT_ID values found in upload: "
            + ", ".join(dup_ids)
            + ". PRODUCT_ID must be unique per product in the file."
        )

    # d) finally: convert PRODUCT_ID to numeric for insert
    df["PRODUCT_ID"] = pd.to_numeric(valid_pid_str, errors="coerce")

    # ------------------------------------------------------------------
    # 7) Clean / validate CARRIER_UPC
    # ------------------------------------------------------------------
    df["CARRIER_UPC"] = df["CARRIER_UPC"].apply(_clean_upc)

    # Missing / blank after cleaning
    missing_upc_mask = df["CARRIER_UPC"].isna() | (df["CARRIER_UPC"] == "")
    if missing_upc_mask.any():
        idxs = df.index[missing_upc_mask].tolist()
        human_rows = [i + 2 for i in idxs]
        errors.append(
            "CARRIER_UPC is required and cannot be blank (or non-numeric) for row(s): "
            + ", ".join(map(str, human_rows))
            + "."
        )

    # Length > 20
    len_mask = (
        df["CARRIER_UPC"].notna()
        & (df["CARRIER_UPC"] != "")
        & (df["CARRIER_UPC"].str.len() > 20)
    )
    if len_mask.any():
        idxs = df.index[len_mask].tolist()
        human_rows = [i + 2 for i in idxs]
        errors.append(
            "CARRIER_UPC must be at most 20 digits for row(s): "
            + ", ".join(map(str, human_rows))
            + "."
        )

    # Reject legacy placeholder
    placeholder_mask = df["CARRIER_UPC"] == "999999999999"
    if placeholder_mask.any():
        idxs = df.index[placeholder_mask].tolist()
        human_rows = [i + 2 for i in idxs]
        errors.append(
            "Placeholder CARRIER_UPC '999999999999' is not allowed for row(s): "
            + ", ".join(map(str, human_rows))
            + ". Please use real UPCs."
        )

    return df, errors


# =============================================================================
# 📥 PRODUCTS UPLOAD
# =============================================================================


def write_products_to_snowflake(df: pd.DataFrame):
    """
    Uploads cleaned products data to Snowflake with proper type handling and auditing.
    """
    df = df.copy()
    df.replace("NAN", np.nan, inplace=True)
    df.replace({np.nan: None}, inplace=True)
    df["CARRIER_UPC"] = df["CARRIER_UPC"].astype(str).str.strip()

    conn, cursor, tenant_id = _get_conn_and_cursor()
    if not conn:
        return

    now_ts, today_date = _build_audit_fields()

    try:
        cursor.execute("BEGIN;")
        cursor.execute("TRUNCATE TABLE PRODUCTS;")

        records = [
            (
                row.PRODUCT_ID,
                row.SUPPLIER,
                row.PRODUCT_NAME,
                row.PACKAGE,
                row.CARRIER_UPC,
                row.PRODUCT_MANAGER,
                tenant_id,
                now_ts,
                now_ts,
                today_date,
            )
            for row in df.itertuples(index=False)
        ]

        insert_sql = """
            INSERT INTO PRODUCTS (
                PRODUCT_ID,
                SUPPLIER,
                PRODUCT_NAME,
                PACKAGE,
                CARRIER_UPC,
                PRODUCT_MANAGER,
                TENANT_ID,
                CREATED_AT,
                UPDATED_AT,
                LAST_LOAD_DATE
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        cursor.executemany(insert_sql, records)
        _finalize_transaction(cursor, conn, "✅ Products uploaded.")

    except Exception as e:
        _rollback_transaction(conn, cursor)
        st.error(f"❌ Product upload failed: {str(e)}")


# =============================================================================
# 📥 SUPPLIER_COUNTY UPLOAD
# =============================================================================


def write_supplier_by_county_to_snowflake(df: pd.DataFrame):
    """
    Uploads formatted Supplier by County DataFrame to SUPPLIER_COUNTY in Snowflake.

    - Tenant-aware connection
    - TRUNCATE + INSERT in a transaction
    - Adds audit fields: TENANT_ID, CREATED_AT, UPDATED_AT, LAST_LOAD_DATE

    Expects columns:
        SUPPLIER, COUNTY, STATUS
    (Case-insensitive; we normalize to upper internally.)
    """
    toml_info = st.session_state.get("toml_info")
    tenant_id = st.session_state.get("tenant_id")

    if not toml_info or not tenant_id:
        st.error("❌ Tenant configuration is missing.")
        return

    # Normalize column names so BOTH legacy and validator flows work:
    # - validator produces SUPPLIER/COUNTY/STATUS
    # - legacy formatter produces Supplier/County/Status
    df = df.copy()
    df.columns = [str(c).strip().upper() for c in df.columns]

    required_cols = ["SUPPLIER", "COUNTY", "STATUS"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(
            "❌ Supplier by County upload failed: missing required columns: "
            + ", ".join(missing)
        )
        return

    # Restrict to required columns only
    df = df[required_cols].copy()

    try:
        conn = connect_to_tenant_snowflake(toml_info)
        if not conn:
            st.error("❌ Failed to connect to Snowflake.")
            return

        cursor = conn.cursor()
        now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today_date = datetime.today().strftime("%Y-%m-%d")

        cursor.execute("BEGIN;")
        cursor.execute("TRUNCATE TABLE SUPPLIER_COUNTY;")

        # Build records using positional tuples – no attribute-name issues
        records = [
            (
                row[0],  # SUPPLIER
                row[1],  # COUNTY
                row[2],  # STATUS
                tenant_id,
                now_ts,
                now_ts,
                today_date,
            )
            for row in df.itertuples(index=False, name=None)
        ]

        insert_sql = """
            INSERT INTO SUPPLIER_COUNTY (
                SUPPLIER,
                COUNTY,
                STATUS,
                TENANT_ID,
                CREATED_AT,
                UPDATED_AT,
                LAST_LOAD_DATE
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s);
        """

        cursor.executemany(insert_sql, records)
        conn.commit()
        st.success("✅ Supplier by County data uploaded successfully to Snowflake.")

    except Exception as e:
        if "conn" in locals():
            conn.rollback()
        st.error(f"❌ Supplier by County upload failed: {str(e)}")

    finally:
        if "cursor" in locals():
            cursor.close()
        if "conn" in locals():
            conn.close()
