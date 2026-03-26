# ---------------- reset_schedule_sections.py ----------------

import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, date, time
from io import BytesIO

from utils.reset_schedule_helpers import (
    format_reset_schedule,
    upload_reset_data,
    generate_reset_schedule_template,
)
from utils.ui_helpers import download_workbook
from utils.snowflake_utils import fetch_distinct_values
from sf_connector.service_connector import connect_to_tenant_snowflake


# ─────────────────────────────────────────────────────────────────────────────
# Section 1: Download Template
# ─────────────────────────────────────────────────────────────────────────────

def render_reset_schedule_formatter_section():
    """
    Step 1: Download template + validate/format uploaded file.
    Fix (2026-03-25): Wrapped file_uploader in st.form to prevent CP reruns
    from dropping the uploaded file before it can be read.
    """
    st.markdown("Download the Reset Schedule template, fill it in, then upload it here for validation and formatting.")

    # Download template (must be outside form — download buttons can't be inside forms)
    tmpl_wb = generate_reset_schedule_template()
    tmpl_buffer = BytesIO()
    tmpl_wb.save(tmpl_buffer)
    tmpl_buffer.seek(0)

    st.download_button(
        label="📥 Download Reset Schedule Template (XLSX)",
        data=tmpl_buffer,
        file_name="Reset_Schedule_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="reset_schedule_template_download",
    )

    st.markdown(
        """
        <small>
        Required columns per row:<br>
        • <b>CHAIN_NAME</b> &nbsp;•&nbsp; <b>STORE_NUMBER</b> &nbsp;•&nbsp; <b>STORE_NAME</b><br>
        • <b>ADDRESS</b> &nbsp;•&nbsp; <b>CITY</b><br>
        • <b>RESET_DATE</b> (mm/dd/yyyy or Excel date)<br>
        • <b>RESET_TIME</b> (e.g. '8:00 AM' or '13:00')<br>
        </small>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("---")
    st.markdown("**Upload completed template for validation & formatting:**")

    with st.form("reset_schedule_formatter_form"):
        uploaded_file = st.file_uploader(
            "Upload Reset Schedule Excel (based on the template)",
            type=["xlsx"],
            key="reset_schedule_upload",
        )
        submitted = st.form_submit_button("Validate & Format")

    if not submitted:
        return

    if uploaded_file is None:
        st.warning("Please upload a Reset Schedule Excel file before submitting.")
        return

    try:
        workbook = openpyxl.load_workbook(uploaded_file)
        with st.spinner("Validating and formatting reset schedule..."):
            formatted_wb = format_reset_schedule(workbook)

        if formatted_wb:
            st.success("✅ Formatting complete. Download to review before upload.")
            download_workbook(formatted_wb, "Formatted_Reset_Schedule.xlsx")
    except Exception as e:
        st.error(f"Failed to format reset schedule: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Section 2: Upload to Database
# ─────────────────────────────────────────────────────────────────────────────

def render_reset_schedule_uploader_section():
    """
    Upload formatted reset schedule to Snowflake.
    Fix (2026-03-25): Wrapped in st.form to prevent CP reruns from dropping
    the uploaded file. Added chain name mismatch validation before upload.
    """
    conn = st.session_state.get("conn")
    if not conn:
        st.error("Database connection not found.")
        return

    try:
        chain_options = fetch_distinct_values(conn, "CUSTOMERS", "CHAIN_NAME")
        chain_options.sort()
    except Exception as e:
        st.error(f"Could not load chain names: {e}")
        return

    st.markdown("Select the chain and upload the formatted Reset Schedule file to push it to the database.")

    with st.form("reset_schedule_uploader_form"):
        selected_chain = st.selectbox(
            "Select Chain Name",
            chain_options,
            key="reset_schedule_chain_select",
        )
        uploaded_file = st.file_uploader(
            "Upload Formatted Reset Schedule File",
            type=["xlsx"],
            key="reset_schedule_final_upload",
        )
        submitted = st.form_submit_button("Upload Reset Schedule to Database")

    if not submitted:
        return

    if uploaded_file is None:
        st.error("Please upload a formatted Reset Schedule file.")
        return

    if not selected_chain:
        st.error("Please select a chain.")
        return

    try:
        df = pd.read_excel(uploaded_file, engine="openpyxl")
        st.markdown("**Preview of uploaded data:**")
        st.dataframe(df.head(), use_container_width=True)

        # Chain name validation — block upload if file doesn't match selection
        if "CHAIN_NAME" in df.columns:
            file_chains = df["CHAIN_NAME"].dropna().str.strip().str.upper().unique().tolist()
            selected_chain_upper = selected_chain.strip().upper()
            mismatched = [c for c in file_chains if c != selected_chain_upper]
            if mismatched:
                st.error(
                    f"❌ Chain name mismatch! You selected **{selected_chain_upper}** "
                    f"but the file contains: **{', '.join(mismatched)}**\n\n"
                    "Please select the correct chain from the dropdown and re-submit."
                )
                return

        with st.spinner("Uploading to database..."):
            upload_reset_data(df, selected_chain)

    except Exception as e:
        st.error(f"❌ Failed to upload reset schedule: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Section 3: Inline Editor (admin only)
# ─────────────────────────────────────────────────────────────────────────────

def render_reset_schedule_editor_section():
    """
    Admin-only inline editor for RESET_DATE and RESET_TIME fields.
    Uses st.data_editor to allow in-table editing without a full re-upload.
    Only changed rows are written back to Snowflake via UPDATE statements.
    """
    if not st.session_state.get("is_admin"):
        return

    conn = st.session_state.get("conn")
    toml_info = st.session_state.get("toml_info")
    if not conn or not toml_info:
        st.error("Database connection not found.")
        return

    st.markdown("Select a chain to view and edit its reset schedule dates and times inline.")

    try:
        chain_options = fetch_distinct_values(conn, "CUSTOMERS", "CHAIN_NAME")
        chain_options.sort()
    except Exception as e:
        st.error(f"Could not load chain names: {e}")
        return

    selected_chain = st.selectbox(
        "Select Chain",
        chain_options,
        key="rs_editor_chain_select",
    )

    if not selected_chain:
        return

    # Load reset schedule for selected chain
    try:
        query = """
            SELECT
                RESET_SCHEDULE_ID,
                STORE_NUMBER,
                STORE_NAME,
                CITY,
                ADDRESS,
                RESET_DATE,
                RESET_TIME,
                STATUS,
                NOTES
            FROM RESET_SCHEDULE
            WHERE UPPER(TRIM(CHAIN_NAME)) = %s
            ORDER BY STORE_NUMBER, RESET_DATE
        """
        fresh_conn = connect_to_tenant_snowflake(toml_info)
        df = pd.read_sql(query, fresh_conn, params=(selected_chain.strip().upper(),))
        fresh_conn.close()
    except Exception as e:
        st.error(f"Failed to load reset schedule: {e}")
        return

    if df.empty:
        st.info(f"No reset schedule records found for {selected_chain}.")
        return

    st.markdown(f"**{len(df)} records** for **{selected_chain}**. Edit RESET_DATE or RESET_TIME inline, then click Save Changes.")

    # Store original for diff comparison
    original_df = df.copy()

    # Configure columns — only RESET_DATE and RESET_TIME are editable
    column_config = {
        "RESET_SCHEDULE_ID": st.column_config.NumberColumn("ID", disabled=True),
        "STORE_NUMBER":       st.column_config.NumberColumn("Store #", disabled=True),
        "STORE_NAME":         st.column_config.TextColumn("Store Name", disabled=True),
        "CITY":               st.column_config.TextColumn("City", disabled=True),
        "ADDRESS":            st.column_config.TextColumn("Address", disabled=True),
        "STATUS":             st.column_config.TextColumn("Status", disabled=True),
        "NOTES":              st.column_config.TextColumn("Notes", disabled=True),
        "RESET_DATE": st.column_config.DateColumn(
            "Reset Date",
            format="MM/DD/YYYY",
            help="Click to change the reset date",
        ),
        "RESET_TIME": st.column_config.TimeColumn(
            "Reset Time",
            format="hh:mm a",
            help="Click to change the reset time",
        ),
    }

    edited_df = st.data_editor(
        df,
        column_config=column_config,
        use_container_width=True,
        hide_index=True,
        key="rs_inline_editor",
        num_rows="fixed",  # prevent adding/deleting rows
    )

    # Find changed rows by comparing RESET_DATE and RESET_TIME only
    changed_mask = (
        (edited_df["RESET_DATE"].astype(str) != original_df["RESET_DATE"].astype(str)) |
        (edited_df["RESET_TIME"].astype(str) != original_df["RESET_TIME"].astype(str))
    )
    changed_df = edited_df[changed_mask]

    if changed_df.empty:
        st.caption("No changes detected.")
    else:
        st.info(f"**{len(changed_df)} row(s)** modified. Click Save Changes to write to the database.")

    if st.button("💾 Save Changes", type="primary", disabled=changed_df.empty):
        if changed_df.empty:
            st.warning("No changes to save.")
            return

        try:
            save_conn = connect_to_tenant_snowflake(toml_info)
            updated = 0
            errors = []

            with save_conn.cursor() as cur:
                for _, row in changed_df.iterrows():
                    try:
                        # Normalize date and time to Snowflake-safe strings
                        reset_date = (
                            row["RESET_DATE"].strftime("%Y-%m-%d")
                            if pd.notna(row["RESET_DATE"]) and hasattr(row["RESET_DATE"], "strftime")
                            else str(row["RESET_DATE"]) if pd.notna(row["RESET_DATE"])
                            else None
                        )
                        reset_time = (
                            row["RESET_TIME"].strftime("%H:%M:%S")
                            if pd.notna(row["RESET_TIME"]) and hasattr(row["RESET_TIME"], "strftime")
                            else str(row["RESET_TIME"]) if pd.notna(row["RESET_TIME"])
                            else None
                        )

                        cur.execute(
                            """
                            UPDATE RESET_SCHEDULE
                            SET RESET_DATE = %s,
                                RESET_TIME = %s,
                                UPDATED_AT = CURRENT_TIMESTAMP()
                            WHERE RESET_SCHEDULE_ID = %s
                            """,
                            (reset_date, reset_time, int(row["RESET_SCHEDULE_ID"])),
                        )
                        updated += 1
                    except Exception as row_err:
                        errors.append(f"Store {row['STORE_NUMBER']}: {row_err}")

                save_conn.commit()
            save_conn.close()

            if errors:
                st.warning(f"Saved {updated} row(s) with {len(errors)} error(s):")
                for err in errors:
                    st.markdown(f"- {err}")
            else:
                st.success(f"✅ {updated} record(s) updated successfully.")
                # Clear cached editor state so table reloads with fresh data
                st.session_state.pop("rs_inline_editor", None)
                st.rerun()

        except Exception as e:
            st.error(f"❌ Save failed: {e}")
